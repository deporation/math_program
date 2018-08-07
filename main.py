# main.py
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from change import change
from datetime import datetime,time
import xlwt
import numpy as np
from scipy.optimize import linprog
import matplotlib.pyplot as plt
ch = []
X = []
Y = []
tr = []
truely = []
number = []
faluse  = []
time_begin = datetime.now()
time_end = datetime.now()
time = datetime.now()
datetime.replace(time,2016,8,1,9,0,0)
workbook = load_workbook('附件一交易数据.xlsx')
booksheet = workbook.active
for i in range(4,37313):
    ch.append(change(booksheet.cell(row = i,column = 1).value,booksheet.cell(row = i,column = 2).value,booksheet.cell(row = i,column = 4).value,booksheet.cell(row = i,column = 5).value,booksheet.cell(row = i,column = 6).value,booksheet.cell(row = i,column = 7).value,booksheet.cell(row = i,column = 8).value,booksheet.cell(row = i,column = 9).value))
    if i ==37312:
        print("successful!")
if i != 37312:
    print("faluse!")
copy = ch
methoad = '卖'
for i in copy:
    if  i.methoad == methoad and i.change_count != 0:
        X.append(i.val)
        Y.append((i.date_detail- time).seconds)
        tr.append(i.trade_number)
    if ((i.date_detail-time).seconds)/(60*60) == 6:
        break
z1 = np.polyfit(X, Y, 6)
p1 = np.poly1d(z1)
plt.plot(X,p1(X),"b-",linewidth=1)
plt.show()
print (z1)
print (p1)
for i in copy:
    for j in tr:
        if i.trade_number ==j :
            truely.append(i)
            number.append(i.val)
    if ((i.date_detail - time).seconds) / (60 * 60) == 6:
        break
print('successful!')
for i in range(len(X)):
    if abs(Y[i] - p1(X[i]))> 5:
        faluse.append(copy[i])
for i in faluse:
    print("异常操作人员：%d",i.trade_number)